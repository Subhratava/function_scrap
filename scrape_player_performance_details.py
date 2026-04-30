import argparse
import asyncio
import json
import re
import sys
import time
import uuid
from pathlib import Path
from urllib.parse import urlparse
from urllib.parse import urlencode

import aiohttp
import pandas as pd
from crawlee import Request
from crawlee.crawlers import ParselCrawler
from crawlee.storage_clients import MemoryStorageClient
from openpyxl import load_workbook


BASE_URL = "https://www.transfermarkt.com"
DEFAULT_INPUT = Path("..") / "player_profiles.xlsx"
DEFAULT_OUTPUT = Path("outputs") / "player_performance_details.xlsx"
DEFAULT_RAW_JSON = Path("outputs") / "player_performance_details.json"
DEFAULT_BACKUP_EVERY = 25

STAT_COLUMNS = [
    "in_squad",
    "appearances",
    "ppg",
    "goals",
    "assists",
    "own_goals",
    "yellow_cards",
    "second_yellow_cards",
    "red_cards",
    "substitutions_on",
    "substitutions_off",
    "penalty_goals",
    "minutes_per_goal",
    "minutes_played",
]
INCLUDED_COMPETITION_TYPE_IDS = {
    1,
    2,
    3,
    4,
    5,
    6,
    7,
    8,
    9,
    10,
    12,
    13,
    14,
    15,
    16,
    18,
    21,
    22,
    23,
    24,
}
TMAPI_BASE_URL = "https://tmapi-alpha.transfermarkt.technology"
_FALLBACK_CRAWLER_LOCK: asyncio.Lock | None = None


def fallback_crawler_lock() -> asyncio.Lock:
    global _FALLBACK_CRAWLER_LOCK
    if _FALLBACK_CRAWLER_LOCK is None:
        _FALLBACK_CRAWLER_LOCK = asyncio.Lock()
    return _FALLBACK_CRAWLER_LOCK


def clean_text(value: str | None) -> str | None:
    if value is None:
        return None
    return " ".join(str(value).replace("\xa0", " ").split()).strip() or None


def selector_text(selector) -> str | None:
    return clean_text(" ".join(selector.xpath(".//text()").getall()))


def absolute_url(href: str | None) -> str | None:
    href = clean_text(href)
    if not href:
        return None
    if href.startswith("http"):
        return href
    return f"{BASE_URL}{href}"


def player_slug_from_href(href: str | None) -> str | None:
    if not href:
        return None
    parts = [part for part in urlparse(href).path.split("/") if part]
    return parts[0] if parts else None


def player_id_from_href(href: str | None) -> str | None:
    if not href:
        return None
    parts = [part for part in urlparse(href).path.rstrip("/").split("/") if part]
    if "spieler" in parts:
        index = parts.index("spieler")
        if index + 1 < len(parts) and parts[index + 1].isdigit():
            return parts[index + 1]
    for part in reversed(parts):
        if part.isdigit():
            return part
    return None


def performance_url_from_href(href: str) -> str | None:
    slug = player_slug_from_href(href)
    player_id = player_id_from_href(href)
    if not slug or not player_id:
        return None
    return f"{BASE_URL}/{slug}/leistungsdatendetails/spieler/{player_id}/plus/1"


def player_from_url(url: str, input_player_name: str | None = None) -> dict:
    path = urlparse(url).path
    slug = player_slug_from_href(path)
    player_id = player_id_from_href(path)
    player_href = f"/{slug}/profil/spieler/{player_id}" if slug and player_id else path
    return {
        "input_row": None,
        "player_id": player_id,
        "input_player_name": input_player_name,
        "player_name": input_player_name,
        "player_href": player_href,
        "performance_url": performance_url_from_href(path) or url,
        "debug_html": None,
    }


def as_int(value: str | None) -> int | None:
    value = clean_text(value)
    if not value or value == "-":
        return None
    value = value.replace("'", "").replace(",", "").replace(".", "")
    match = re.search(r"-?\d+", value)
    if not match:
        return None
    return int(match.group(0))


def as_float(value: str | None) -> float | None:
    value = clean_text(value)
    if not value or value == "-":
        return None
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return None


def read_players(input_file: Path, sheet_name: str, player_href_column: str) -> list[dict]:
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header = [clean_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_lookup = {name: index for index, name in enumerate(header) if name}
    if player_href_column not in header_lookup:
        raise ValueError(f"Could not find column '{player_href_column}' in {input_file}")

    href_idx = header_lookup[player_href_column]
    name_idx = header_lookup.get("input_player_name")
    player_name_idx = header_lookup.get("player_name")
    id_idx = header_lookup.get("player_id")

    seen = set()
    players = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        href = clean_text(row[href_idx]) if href_idx < len(row) else None
        if not href:
            continue

        player_id = clean_text(row[id_idx]) if id_idx is not None and id_idx < len(row) else player_id_from_href(href)
        key = player_id or href
        if key in seen:
            continue
        seen.add(key)

        input_name = clean_text(row[name_idx]) if name_idx is not None and name_idx < len(row) else None
        player_name = clean_text(row[player_name_idx]) if player_name_idx is not None and player_name_idx < len(row) else None
        players.append(
            {
                "input_row": row_number,
                "player_id": player_id or player_id_from_href(href),
                "input_player_name": input_name or player_name,
                "player_name": player_name,
                "player_href": href,
                "performance_url": performance_url_from_href(href),
            }
        )

    return players


def parse_stat_value(column: str, value: str | None):
    if column == "ppg":
        return as_float(value)
    return as_int(value)


async def fetch_json(url: str) -> dict:
    timeout = aiohttp.ClientTimeout(total=30)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
    }
    async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
        async with session.get(url) as response:
            status = response.status
            body = await response.text(errors="replace")

    if body is None:
        raise RuntimeError(f"No response body fetched from {url}")
    if not str(body).lstrip().startswith("{"):
        raise RuntimeError(f"Expected JSON from {url}, got HTTP {status}: {body[:200]}")
    return json.loads(body)


async def fetch_performance_payload(player: dict) -> dict:
    player_id = player["player_id"]
    if not player_id:
        raise ValueError("Could not build performance API URL without player_id")
    body = await fetch_json(f"{BASE_URL}/ceapi/performance-game/{player_id}")
    data = body.get("data")
    if not data:
        raise ValueError(body.get("message") or "Performance API returned no data")
    return data


async def fetch_competition_lookup(competition_ids: list[str]) -> dict:
    ids = [competition_id for competition_id in dict.fromkeys(competition_ids) if competition_id]
    if not ids:
        return {}
    query = urlencode({"ids[]": ids}, doseq=True)
    body = await fetch_json(f"{TMAPI_BASE_URL}/competitions?{query}")
    return {str(item.get("id")): item for item in body.get("data") or []}


def add_api_stat_text(row: dict) -> dict:
    for column in STAT_COLUMNS:
        value = row.get(column)
        if column in {"minutes_per_goal", "minutes_played"} and isinstance(value, int):
            row[f"{column}_text"] = f"{value:,}'"
        elif column == "ppg" and value is not None:
            row[f"{column}_text"] = f"{value:.2f}"
        elif value is None:
            row[f"{column}_text"] = "-"
        else:
            row[f"{column}_text"] = str(value)
    return row


def increment_counter(row: dict, key: str, value: int | float | None) -> None:
    row[key] = row.get(key, 0) + (value or 0)


def build_performance_rows_from_api(player: dict, payload: dict, competitions: dict) -> list[dict]:
    grouped = {}
    performance_rows = payload.get("performance") or []

    for performance in performance_rows:
        game = performance.get("gameInformation") or {}
        if game.get("competitionTypeId") not in INCLUDED_COMPETITION_TYPE_IDS:
            continue

        stats = performance.get("statistics") or {}
        general = stats.get("generalStatistics") or {}
        goals = stats.get("goalStatistics") or {}
        cards = stats.get("cardStatistics") or {}
        playing_time = stats.get("playingTimeStatistics") or {}
        clubs = performance.get("clubsInformation") or {}
        club = clubs.get("club") or {}
        participation_state = general.get("participationState")
        if participation_state not in {"played", "in squad"}:
            continue

        competition_id = str(game.get("competitionId") or "")
        season_id = game.get("seasonId")
        club_id = str(club.get("clubId") or "")
        key = (season_id, competition_id, club_id)
        competition = competitions.get(competition_id) or {}
        season = game.get("season") or {}

        if key not in grouped:
            competition_href = competition.get("relativeUrl")
            grouped[key] = {
                "player_id": player["player_id"],
                "input_player_name": player["input_player_name"],
                "player_name": player["player_name"],
                "player_href": player["player_href"],
                "performance_url": player["performance_url"],
                "row_type": "competition",
                "row_index": None,
                "season_id": season_id,
                "season": season.get("display"),
                "club_id": club_id,
                "competition_id": competition_id,
                "competition": competition.get("name") or competition_id,
                "competition_href": competition_href,
                "competition_url": absolute_url(competition_href),
                "points": 0,
                "in_squad": 0,
                "appearances": 0,
                "goals": 0,
                "assists": 0,
                "own_goals": 0,
                "yellow_cards": 0,
                "second_yellow_cards": 0,
                "red_cards": 0,
                "substitutions_on": 0,
                "substitutions_off": 0,
                "penalty_goals": 0,
                "minutes_played": 0,
                "status": "ok",
                "error": None,
            }

        row = grouped[key]
        row["in_squad"] += 1
        if participation_state == "played":
            row["appearances"] += 1
        increment_counter(row, "points", general.get("pointsOnThePitch"))
        increment_counter(row, "goals", goals.get("goalsScoredTotal"))
        increment_counter(row, "assists", goals.get("assists"))
        increment_counter(row, "own_goals", goals.get("ownGoalsScored"))
        increment_counter(row, "yellow_cards", cards.get("yellowCardNet") or cards.get("yellowCardGross"))
        row["second_yellow_cards"] += 1 if cards.get("yellowRedCard") else 0
        row["red_cards"] += 1 if cards.get("redCard") else 0
        row["substitutions_on"] += 1 if playing_time.get("substitutedIn") else 0
        row["substitutions_off"] += 1 if playing_time.get("substitutedOut") else 0
        increment_counter(row, "penalty_goals", goals.get("penaltyShooterGoalsScored"))
        increment_counter(row, "minutes_played", playing_time.get("playedMinutes"))

    rows = []
    for row_index, row in enumerate(grouped.values(), start=1):
        row["row_index"] = row_index
        row["ppg"] = round(row["points"] / row["appearances"], 2) if row["appearances"] else None
        row["minutes_per_goal"] = round(row["minutes_played"] / row["goals"]) if row["goals"] else None
        rows.append(add_api_stat_text(row))

    if rows:
        total = {
            "player_id": player["player_id"],
            "input_player_name": player["input_player_name"],
            "player_name": player["player_name"],
            "player_href": player["player_href"],
            "performance_url": player["performance_url"],
            "row_type": "total",
            "row_index": len(rows) + 1,
            "season_id": None,
            "season": None,
            "club_id": None,
            "competition_id": None,
            "competition": "Total",
            "competition_href": None,
            "competition_url": None,
            "points": sum(row.get("points") or 0 for row in rows),
            "status": "ok",
            "error": None,
        }
        for column in STAT_COLUMNS:
            if column in {"ppg", "minutes_per_goal"}:
                continue
            total[column] = sum(row.get(column) or 0 for row in rows)
        total["ppg"] = round(total["points"] / total["appearances"], 2) if total["appearances"] else None
        total["minutes_per_goal"] = round(total["minutes_played"] / total["goals"]) if total["goals"] else None
        rows.append(add_api_stat_text(total))

    if not rows:
        return [failure_row(player, "Performance API returned no included competition rows")]
    return rows


async def scrape_performance_api(player: dict) -> list[dict]:
    try:
        payload = await fetch_performance_payload(player)
        competitions = await fetch_competition_lookup(payload.get("competitionIds") or [])
        return build_performance_rows_from_api(player, payload, competitions)
    except Exception as exc:
        return [failure_row(player, str(exc))]


def page_debug(selector) -> dict:
    return {
        "page_title": clean_text(selector.xpath("string(//title)").get()),
        "page_h1": clean_text(selector.xpath("string(//h1)").get()),
        "items_table_count": len(selector.xpath("//table[contains(concat(' ', normalize-space(@class), ' '), ' items ')]")),
    }


def parse_performance_page(selector, player: dict) -> list[dict]:
    tables = selector.xpath("//table[contains(concat(' ', normalize-space(@class), ' '), ' items ')]")
    table = None
    for candidate in tables:
        first_header = clean_text(candidate.xpath("string(.//thead//th[1])").get())
        first_body_cell = clean_text(candidate.xpath("string(.//tbody/tr[1]/td[1])").get())
        if first_header == "Competition" or first_body_cell:
            table = candidate
            break

    if table is None:
        debug = page_debug(selector)
        return [
            failure_row(
                player,
                (
                    "Could not find the performance details table"
                    f" (title={debug['page_title']!r}, h1={debug['page_h1']!r}, items_tables={debug['items_table_count']})"
                ),
            )
        ]

    rows = []
    data_rows = table.xpath(".//tbody/tr[td] | .//tfoot/tr[td]")
    for row_index, row in enumerate(data_rows, start=1):
        cells = row.xpath("./td")
        if not cells:
            continue

        first_cell_text = selector_text(cells[0])
        row_type = "total" if first_cell_text and first_cell_text.rstrip(":").lower() == "total" else "competition"
        competition_link = cells[0].xpath(".//a[1]/@href").get()
        competition_name = clean_text(cells[0].xpath("string(.//a[1])").get()) or first_cell_text
        if row_type == "total":
            competition_name = "Total"

        stat_cells = cells[1:]
        raw_values = [selector_text(cell) for cell in stat_cells]
        if not any(raw_values) and row_type != "total":
            continue

        row_data = {
            "player_id": player["player_id"],
            "input_player_name": player["input_player_name"],
            "player_name": player["player_name"],
            "player_href": player["player_href"],
            "performance_url": player["performance_url"],
            "row_type": row_type,
            "row_index": row_index,
            "competition": competition_name,
            "competition_href": competition_link,
            "competition_url": absolute_url(competition_link),
            "status": "ok",
            "error": None,
        }

        for column, value in zip(STAT_COLUMNS, raw_values):
            row_data[column] = parse_stat_value(column, value)
            row_data[f"{column}_text"] = clean_text(value)
        for column in STAT_COLUMNS[len(raw_values):]:
            row_data[column] = None
            row_data[f"{column}_text"] = None

        rows.append(row_data)

    if not rows:
        return [failure_row(player, "Performance details table was found but no rows were parsed")]
    return rows


def failure_row(player: dict, error: str | None) -> dict:
    row = {
        "player_id": player["player_id"],
        "input_player_name": player["input_player_name"],
        "player_name": player["player_name"],
        "player_href": player["player_href"],
        "performance_url": player["performance_url"],
        "row_type": None,
        "row_index": None,
        "season_id": None,
        "season": None,
        "club_id": None,
        "competition_id": None,
        "competition": None,
        "competition_href": None,
        "competition_url": None,
        "points": None,
        "status": "failed",
        "error": error,
    }
    for column in STAT_COLUMNS:
        row[column] = None
        row[f"{column}_text"] = None
    return row


async def scrape_performance(player: dict) -> list[dict]:
    if not player["performance_url"]:
        return [failure_row(player, "Could not build performance details URL from player_href")]
    api_rows = await scrape_performance_api(player)
    if api_rows and api_rows[0].get("status") == "ok":
        return api_rows

    async with fallback_crawler_lock():
        crawler = ParselCrawler(storage_client=MemoryStorageClient())
        parsed = {"seen": False, "rows": None}

        @crawler.router.default_handler
        async def parse(context):
            parsed["seen"] = True
            debug_html = player.get("debug_html")
            if debug_html:
                Path(debug_html).parent.mkdir(parents=True, exist_ok=True)
                Path(debug_html).write_text(context.selector.get() or "", encoding="utf-8")
            parsed["rows"] = parse_performance_page(context.selector, player)

        try:
            await crawler.run([Request.from_url(player["performance_url"], unique_key=f"{player['performance_url']}#{uuid.uuid4()}")])
        except Exception as exc:
            return [failure_row(player, str(exc))]

        if parsed["rows"]:
            return parsed["rows"]
        return [failure_row(player, "Crawler finished without parsing the performance details page")]


def save_json(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def save_xlsx(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Performance details")
        worksheet = writer.sheets["Performance details"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def backup_xlsx_path(output_path: Path, backup_path: Path | None) -> Path:
    if backup_path is not None:
        return backup_path
    return output_path.with_name(f"{output_path.stem}_backup{output_path.suffix}")


async def run(args: argparse.Namespace) -> list[dict]:
    if args.url:
        player = player_from_url(args.url, args.input_player_name)
        player["debug_html"] = args.debug_html
        players = [player]
    else:
        players = read_players(args.input, args.sheet, args.player_href_column)
    if args.limit_players:
        players = players[: args.limit_players]

    print(f"Loaded {len(players)} unique players from {args.input}")
    if not players:
        return []

    concurrency = max(1, args.concurrency)
    print(f"Using concurrency={concurrency}, request_delay={args.request_delay}s")

    semaphore = asyncio.Semaphore(concurrency)
    start_lock = asyncio.Lock()
    last_start = 0.0

    async def wait_for_start_slot() -> None:
        nonlocal last_start
        if args.request_delay <= 0:
            return
        async with start_lock:
            now = time.monotonic()
            wait_seconds = max(0.0, last_start + args.request_delay - now)
            if wait_seconds > 0:
                await asyncio.sleep(wait_seconds)
            last_start = time.monotonic()

    async def scrape_player(index: int, player: dict) -> tuple[int, list[dict]]:
        async with semaphore:
            await wait_for_start_slot()
            label = player["input_player_name"] or player["player_href"]
            print(f"[{index}/{len(players)}] {label}")
            try:
                return index, await scrape_performance(player)
            except Exception as exc:
                return index, [failure_row(player, str(exc))]

    def completed_rows_in_input_order(results_by_index: list[list[dict] | None]) -> list[dict]:
        ordered_rows = []
        for player_rows in results_by_index:
            if player_rows:
                ordered_rows.extend(player_rows)
        return ordered_rows

    rows_by_index: list[list[dict] | None] = [None] * len(players)
    tasks = [
        asyncio.create_task(scrape_player(index, player))
        for index, player in enumerate(players, start=1)
    ]
    completed = 0
    for task in asyncio.as_completed(tasks):
        index, player_rows = await task
        rows_by_index[index - 1] = player_rows
        completed += 1
        ok_rows = sum(1 for row in player_rows if row.get("status") == "ok")
        print(f"  [{index}/{len(players)}] performance_rows={ok_rows} status={player_rows[0].get('status')}")
        if args.backup_every and args.backup_every > 0 and completed % args.backup_every == 0:
            rows = completed_rows_in_input_order(rows_by_index)
            backup_path = backup_xlsx_path(args.xlsx, args.backup_xlsx)
            save_xlsx(rows, backup_path)
            print(f"  backup saved after {completed} players to {backup_path}")

    return completed_rows_in_input_order(rows_by_index)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape Transfermarkt detailed performance-by-competition tables from player_profiles.xlsx."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--sheet", default="Player data")
    parser.add_argument("--player-href-column", default="player_href")
    parser.add_argument("--url", default=None, help="Scrape one Transfermarkt detailed performance URL directly.")
    parser.add_argument("--input-player-name", default=None, help="Optional name to store when using --url.")
    parser.add_argument("--debug-html", type=Path, default=None, help="Optional path to save the fetched HTML when using --url.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--json", type=Path, default=DEFAULT_RAW_JSON)
    parser.add_argument("--limit-players", type=int, default=None)
    parser.add_argument(
        "--concurrency",
        type=int,
        default=4,
        help="Maximum number of players to scrape at the same time.",
    )
    parser.add_argument("--request-delay", type=float, default=1.5)
    parser.add_argument(
        "--backup-every",
        type=int,
        default=DEFAULT_BACKUP_EVERY,
        help="Save a backup Excel workbook after this many processed players. Use 0 to disable.",
    )
    parser.add_argument(
        "--backup-xlsx",
        type=Path,
        default=None,
        help="Optional backup Excel path. Defaults to the output path with '_backup' added.",
    )
    return parser.parse_args()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(errors="replace")

    args = parse_args()
    start = time.time()
    rows = asyncio.run(run(args))
    save_json(rows, args.json)
    save_xlsx(rows, args.xlsx)
    print(f"Saved {len(rows)} performance-detail rows to {args.xlsx}")
    print(f"Saved raw JSON to {args.json}")
    print(f"Done in {time.time() - start:.1f}s")


if __name__ == "__main__":
    main()
